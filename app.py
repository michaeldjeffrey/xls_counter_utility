from openpyxl import Workbook, load_workbook
from collections import Counter
import json

from flask import Flask, request, render_template, send_file

app = Flask(__name__)

def index_post(file, **kwargs):
    # Workbook things
    workbook = load_workbook(file)
    sheet = workbook.get_active_sheet()

    # set the variables for reading and writing
    starting_point = kwargs.pop('starting_point')
    starting_point = starting_point[0] if starting_point[0] else 3

    column = kwargs.pop('id_column')
    column = column[0] if column[0] else 'A'

    filename = kwargs.pop('destination_filename')
    filename = filename[0] if filename[0] else 'Student Counter'

    # variables to use, bruh
    folder = 'export/'
    destination_filename = "{}{}.xls".format(folder, filename)
    accetable_null_values = ['-', '=', '+']
    student_attendance = Counter()

    # loop
    going = True
    while going:
        lookup_string = "{}{}".format(column, starting_point)
        cell = sheet[lookup_string].value
        if isinstance(cell, int):
            student_attendance[cell] += 1
            starting_point += 1
        elif cell in accetable_null_values:
            starting_point += 1
        else:
            going = False

    wb = Workbook()
    sheet = wb.active
    index = 1
    for key, value, in student_attendance.items():
        sheet.cell("A{}".format(index)).value = key
        sheet.cell("B{}".format(index)).value = value
        index += 1

    wb.save(filename=destination_filename)

    return send_file(destination_filename,
    		attachment_filename=destination_filename,
    		as_attachment=True,
    		mimetype='application/vnd.ms-excel')


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files['file']
        return index_post(file, **request.form)
    else:
        return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True, port=8080)
